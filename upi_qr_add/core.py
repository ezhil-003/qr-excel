from __future__ import annotations

import shutil
from dataclasses import dataclass
from enum import Enum
from pathlib import Path
from typing import Any
from uuid import uuid4

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from tqdm import tqdm

from .logger import SQLiteLogger
from .qr_generator import create_decorated_qr_image
from .utils import (
    build_upi_deep_link,
    checkpoint_key,
    detect_amount_header,
    find_header_index,
    output_excel_path,
    parse_amount,
)


class QRMode(str, Enum):
    EMBED = "embed"
    HYPERLINK = "hyperlink"


@dataclass(slots=True)
class ProcessConfig:
    vpa: str
    payee_name: str
    note: str = "Payment for order"
    mode: QRMode = QRMode.EMBED
    logo_path: Path | None = None
    db_path: Path = Path("upi_qr_log.db")


@dataclass(slots=True)
class ProcessSummary:
    total_rows: int
    successful: int
    failed: int
    skipped: int
    output_file: Path
    resumed_from: int | None
    interrupted: bool
    status: str
    session_id: str
    error_message: str | None = None

    @property
    def failed_total(self) -> int:
        return self.failed + self.skipped


def _cleanup_temp_files(temp_files: list[Path], temp_dir: Path) -> None:
    for file in temp_files:
        try:
            file.unlink(missing_ok=True)
        except OSError:
            continue
    try:
        temp_dir.rmdir()
    except OSError:
        pass


def process_workbook(
    input_file: str | Path,
    config: ProcessConfig,
    *,
    stop_after_rows: int | None = None,
) -> ProcessSummary:
    input_path = Path(input_file).expanduser().resolve()
    db_path = Path(config.db_path).expanduser().resolve()
    output_path = output_excel_path(input_path)

    successful = 0
    failed = 0
    skipped = 0
    interrupted = False
    resumed_from: int | None = None
    total_rows = 0
    wb = None
    ws = None
    temp_files: list[Path] = []
    temp_dir: Path | None = None
    run_key = ""

    with SQLiteLogger(db_path) as logger:
        logger.set_session_state(status="running", input_file=input_path, output_file=output_path)
        logger.log_session_event(step="run_started", status="success")

        try:
            if not input_path.exists():
                raise FileNotFoundError(f"Input file not found: {input_path}")
            if input_path.suffix.lower() != ".xlsx":
                raise ValueError("Input must be an .xlsx file.")

            if output_path.exists():
                workbook_path = output_path
                copied_fresh = False
            else:
                shutil.copy2(input_path, output_path)
                workbook_path = output_path
                copied_fresh = True

            wb = load_workbook(workbook_path)
            if not wb.worksheets:
                raise ValueError("Workbook has no worksheets.")
            ws = wb.worksheets[0]

            amount_header = detect_amount_header(ws)
            if amount_header is None:
                raise ValueError(
                    'Missing required amount column. Expected header like "Amount" '
                    'or "Balance_Amount(Rs)" within first 25 rows.'
                )
            header_row, amount_col, _matched_header = amount_header

            qr_col = find_header_index(ws, "payment_qr", row=header_row)
            if qr_col is None:
                qr_col = ws.max_column + 1
                ws.cell(row=header_row, column=qr_col, value="payment_qr")

            qr_col_letter = get_column_letter(qr_col)
            ws.column_dimensions[qr_col_letter].width = max(
                ws.column_dimensions[qr_col_letter].width or 0,
                24,
            )

            run_key = checkpoint_key(input_path, output_path, ws.title)
            last_successful_row = logger.get_checkpoint(run_key)
            if copied_fresh and last_successful_row > 1:
                logger.reset_checkpoint(run_key)
                last_successful_row = 1

            resumed_from = last_successful_row if last_successful_row > 1 else None
            start_row = max(header_row + 1, last_successful_row + 1)
            total_rows = max(ws.max_row - header_row, 0)

            temp_dir = output_path.parent / f".{output_path.stem}_tmp_qr"
            hyper_dir = output_path.parent / f"{output_path.stem}_qr_images"

            if config.mode == QRMode.EMBED:
                temp_dir.mkdir(parents=True, exist_ok=True)
            else:
                hyper_dir.mkdir(parents=True, exist_ok=True)

            rows_to_process = list(range(start_row, ws.max_row + 1))
            progress = tqdm(
                rows_to_process,
                total=len(rows_to_process),
                desc="Processing rows",
                unit="row",
            )

            try:
                for row_idx in progress:
                    raw_amount: Any = ws.cell(row=row_idx, column=amount_col).value
                    parsed_amount = parse_amount(raw_amount)
                    if parsed_amount is None or parsed_amount <= 0:
                        skipped += 1
                        logger.log_step(
                            row_index=row_idx,
                            amount=None,
                            txn_id=None,
                            step="validate_amount",
                            status="failed",
                            error_type="InvalidAmount",
                            error_message="Amount is missing, non-numeric, zero, or negative.",
                        )
                        continue

                    txn_id = str(uuid4())
                    current_step = "generate_upi_url"
                    try:
                        upi_url = build_upi_deep_link(
                            vpa=config.vpa,
                            payee_name=config.payee_name,
                            amount=parsed_amount,
                            txn_id=txn_id,
                            note=config.note,
                        )
                        logger.log_step(
                            row_index=row_idx,
                            amount=parsed_amount,
                            txn_id=txn_id,
                            step=current_step,
                            status="success",
                        )

                        current_step = "create_qr_with_logo"
                        qr_image = create_decorated_qr_image(
                            upi_url,
                            logo_path=config.logo_path,
                        )
                        logger.log_step(
                            row_index=row_idx,
                            amount=parsed_amount,
                            txn_id=txn_id,
                            step=current_step,
                            status="success",
                        )

                        target_cell = f"{qr_col_letter}{row_idx}"
                        if config.mode == QRMode.EMBED:
                            current_step = "embed_image"
                            temp_path = temp_dir / f"row_{row_idx}_{txn_id}.png"
                            qr_image.save(temp_path, format="PNG")
                            xl_img = XLImage(str(temp_path))
                            xl_img.width = 120
                            xl_img.height = 120
                            ws.add_image(xl_img, target_cell)
                            ws.row_dimensions[row_idx].height = max(
                                ws.row_dimensions[row_idx].height or 0,
                                96,
                            )
                            temp_files.append(temp_path)
                        else:
                            current_step = "save_image"
                            file_name = f"qr_row_{row_idx}.png"
                            image_path = hyper_dir / file_name
                            qr_image.save(image_path, format="PNG")
                            rel_path = f"{hyper_dir.name}/{file_name}"
                            cell = ws.cell(row=row_idx, column=qr_col, value="Open QR")
                            cell.hyperlink = rel_path
                            cell.style = "Hyperlink"
                            current_step = "add_hyperlink"

                        logger.log_step(
                            row_index=row_idx,
                            amount=parsed_amount,
                            txn_id=txn_id,
                            step=current_step,
                            status="success",
                        )
                        logger.update_checkpoint(run_key, row_idx)
                        successful += 1

                        if stop_after_rows is not None and successful >= stop_after_rows:
                            raise KeyboardInterrupt("Simulated interruption for testing.")

                    except KeyboardInterrupt:
                        raise
                    except Exception as exc:  # pragma: no cover - branch is tested via behavior.
                        failed += 1
                        logger.log_step(
                            row_index=row_idx,
                            amount=parsed_amount,
                            txn_id=txn_id,
                            step=current_step,
                            status="failed",
                            error_type=type(exc).__name__,
                            error_message=str(exc),
                        )
            except KeyboardInterrupt:
                interrupted = True
                logger.log_step(
                    row_index=None,
                    amount=None,
                    txn_id=None,
                    step="processing_interrupted",
                    status="failed",
                    error_type="KeyboardInterrupt",
                    error_message="Execution interrupted. Resume supported on next run.",
                )
                logger.log_session_event(
                    step="processing_interrupted",
                    status="failed",
                    error_type="KeyboardInterrupt",
                    error_message="Execution interrupted. Resume supported on next run.",
                )
            finally:
                progress.close()

            if wb is not None:
                wb.save(output_path)
            if config.mode == QRMode.EMBED and temp_dir is not None:
                _cleanup_temp_files(temp_files, temp_dir)

            if interrupted:
                status = "interrupted"
            elif failed + skipped > 0:
                status = "completed_with_errors"
            else:
                status = "completed"

            logger.set_session_state(status=status, input_file=input_path, output_file=output_path)
            logger.log_summary(
                total_rows=total_rows,
                successful_rows=successful,
                failed_rows=failed,
                skipped_rows=skipped,
                output_file=output_path,
            )
            logger.log_session_event(step="run_finished", status="success")

            return ProcessSummary(
                total_rows=total_rows,
                successful=successful,
                failed=failed,
                skipped=skipped,
                output_file=output_path,
                resumed_from=resumed_from,
                interrupted=interrupted,
                status=status,
                session_id=logger.session_id,
            )

        except Exception as exc:
            logger.log_session_event(
                step="setup_failed",
                status="failed",
                error_type=type(exc).__name__,
                error_message=str(exc),
            )
            logger.set_session_state(
                status="setup_failed",
                input_file=input_path,
                output_file=output_path,
            )
            logger.log_summary(
                total_rows=0,
                successful_rows=0,
                failed_rows=1,
                skipped_rows=0,
                output_file=output_path,
            )
            return ProcessSummary(
                total_rows=0,
                successful=0,
                failed=1,
                skipped=0,
                output_file=output_path,
                resumed_from=None,
                interrupted=False,
                status="setup_failed",
                session_id=logger.session_id,
                error_message=str(exc),
            )
