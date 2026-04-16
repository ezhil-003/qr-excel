from __future__ import annotations

import os
import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Any
from urllib.parse import quote, urlencode

from openpyxl.worksheet.worksheet import Worksheet


def normalize_header(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip().casefold()


def canonical_header(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "", normalize_header(value))


def find_header_index(ws: Worksheet, header_name: str, *, row: int = 1) -> int | None:
    target = header_name.strip().casefold()
    for col in range(1, ws.max_column + 1):
        if normalize_header(ws.cell(row=row, column=col).value) == target:
            return col
    return None


def detect_amount_header(
    ws: Worksheet,
    *,
    max_scan_rows: int = 25,
) -> tuple[int, int, str] | None:
    """
    Detect amount-like column headers in the top rows.

    Returns: (header_row, amount_col, matched_header_text)
    """

    accepted = {
        "amount",
        "balanceamount",
        "balanceamountrs",
    }

    scan_until = min(max_scan_rows, ws.max_row)
    for row_idx in range(1, scan_until + 1):
        for col_idx in range(1, ws.max_column + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            header_text = normalize_header(value)
            if not header_text:
                continue
            if canonical_header(header_text) in accepted:
                return row_idx, col_idx, str(value).strip()
    return None


def parse_amount(value: Any) -> float | None:
    if value is None:
        return None

    if isinstance(value, bool):
        return None

    if isinstance(value, (int, float)):
        return float(value)

    if isinstance(value, str):
        cleaned = value.strip().replace(",", "")
        if not cleaned:
            return None
        try:
            return float(cleaned)
        except ValueError:
            return None

    return None


def build_upi_deep_link(
    *,
    vpa: str,
    payee_name: str,
    amount: float,
    txn_id: str,
    note: str,
) -> str:
    query = urlencode(
        {
            "pa": vpa,
            "pn": payee_name,
            "am": f"{amount:.2f}",
            "tr": txn_id,
            "tn": note,
            "cu": "INR",
        },
        quote_via=quote,
    )
    return f"upi://pay?{query}"


def default_logo_path() -> Path:
    return Path(__file__).resolve().parent / "assets" / "upi_logo.png"


def template_db_path() -> Path:
    return Path(__file__).resolve().parent / "assets" / "upi_qr_template.db"


def app_runtime_dir() -> Path:
    return Path.home() / ".upi-qr-add"


def app_sessions_dir() -> Path:
    return app_runtime_dir() / "sessions"


def make_session_id() -> str:
    now = datetime.now(timezone.utc).strftime("%Y%m%d%H%M%S")
    return f"{now}_{os.getpid()}"


def session_db_path(session_id: str) -> Path:
    return app_sessions_dir() / f"session_{session_id}.db"


def output_excel_path(input_file: Path) -> Path:
    return input_file.with_name(f"{input_file.stem}_with_qr.xlsx")


def checkpoint_key(input_file: Path, output_file: Path, sheet_name: str) -> str:
    return f"{input_file.resolve()}::{output_file.resolve()}::{sheet_name}"
