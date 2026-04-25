"""Excel openpyxl manipulations."""

from __future__ import annotations

from pathlib import Path
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor, XDRPositiveSize2D
from openpyxl.utils.units import pixels_to_EMU

from ..core.exceptions import ExcelProcessingError

QR_COLUMN_MIN_WIDTH = 26
QR_IMAGE_SIZE_PX = 140
QR_ROW_MIN_HEIGHT = 110

def setup_qr_column(ws: Worksheet, qr_col: int, header_row: int = 1) -> str:
    """Set the header and adjust the width for the QR code column."""
    try:
        ws.cell(row=header_row, column=qr_col, value="payment_qr")
        qr_col_letter = get_column_letter(qr_col)
        ws.column_dimensions[qr_col_letter].width = max(
            ws.column_dimensions[qr_col_letter].width or 0,
            QR_COLUMN_MIN_WIDTH,
        )
        return qr_col_letter
    except Exception as e:
        raise ExcelProcessingError(f"Failed to setup QR column: {e}")

def embed_qr_image(ws: Worksheet, image_path: Path, row_idx: int, col_letter: str) -> None:
    """Embed the QR image into the specified cell, centered within the cell."""
    try:
        # 1. Adjust row height FIRST so dimensions are stable for anchoring
        current_height = ws.row_dimensions[row_idx].height or 15
        ws.row_dimensions[row_idx].height = max(current_height, QR_ROW_MIN_HEIGHT)
        new_height = ws.row_dimensions[row_idx].height

        # 2. Get column width and calculate centering offsets
        # Approx: 1 width unit = 7.2 pixels, 1 height point = 1.33 pixels
        col_width = ws.column_dimensions[col_letter].width or 8.43
        cell_w_px = col_width * 7.2
        cell_h_px = new_height * 1.33

        offset_x = max(0, int((cell_w_px - QR_IMAGE_SIZE_PX) / 2))
        offset_y = max(0, int((cell_h_px - QR_IMAGE_SIZE_PX) / 2))

        # 3. Create and configure the image
        xl_img = XLImage(str(image_path))
        xl_img.width = QR_IMAGE_SIZE_PX
        xl_img.height = QR_IMAGE_SIZE_PX

        # 4. Use OneCellAnchor for precise centering
        col_idx = column_index_from_string(col_letter)
        marker = AnchorMarker(
            col=col_idx - 1,
            colOff=pixels_to_EMU(offset_x),
            row=row_idx - 1,
            rowOff=pixels_to_EMU(offset_y)
        )
        # Create the extent using image dimensions
        size = XDRPositiveSize2D(pixels_to_EMU(xl_img.width), pixels_to_EMU(xl_img.height))
        xl_img.anchor = OneCellAnchor(_from=marker, ext=size)
        
        ws.add_image(xl_img)
    except Exception as e:
        raise ExcelProcessingError(f"Failed to embed QR image at row {row_idx}: {e}")

def add_qr_hyperlink(ws: Worksheet, row_idx: int, col_idx: int, rel_path: str) -> None:
    """Add a hyperlink to an external QR image."""
    try:
        cell = ws.cell(row=row_idx, column=col_idx, value="Open QR")
        cell.hyperlink = rel_path
        cell.style = "Hyperlink"
    except Exception as e:
        raise ExcelProcessingError(f"Failed to add QR hyperlink at row {row_idx}: {e}")
