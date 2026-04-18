"""Excel cell parsing and header detection."""

from __future__ import annotations

import re
from typing import Any

from openpyxl.worksheet.worksheet import Worksheet


def normalize_header(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip().casefold()


def canonical_header(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "", normalize_header(value))


def find_header_index(ws: Worksheet, header_name: str, *, row: int = 1) -> int | None:
    target = header_name.strip().casefold()
    for col in range(1, ws.max_column + 1):
        if normalize_header(ws.cell(row=row, column=col).value) == target:
            return col
    return None


def detect_amount_header(
    ws: Worksheet,
    *,
    max_scan_rows: int = 25,
) -> tuple[int, int, str] | None:
    """
    Detect amount-like column headers in the top rows.

    Returns: (header_row, amount_col, matched_header_text)
    """

    accepted = {
        "amount",
        "balanceamount",
        "balanceamountrs",
    }

    scan_until = min(max_scan_rows, ws.max_row)
    for row_idx in range(1, scan_until + 1):
        for col_idx in range(1, ws.max_column + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            header_text = normalize_header(value)
            if not header_text:
                continue
            if canonical_header(header_text) in accepted:
                return row_idx, col_idx, str(value).strip()
    return None


def parse_amount(value: Any) -> float | None:
    if value is None:
        return None

    if isinstance(value, bool):
        return None

    if isinstance(value, (int, float)):
        return float(value)

    if isinstance(value, str):
        cleaned = (
            value.strip()
            .replace(",", "")
            .replace("\u20b9", "")   # ₹ symbol
            .replace("Rs.", "")
            .replace("Rs", "")
            .replace("INR", "")
            .strip()
        )
        if not cleaned:
            return None
        try:
            return float(cleaned)
        except ValueError:
            return None

    return None
