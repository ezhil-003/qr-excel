"""Core processing pipeline: workbook iteration, QR generation, and Excel output."""

from __future__ import annotations

import sqlite3
import shutil
import sys
from pathlib import Path
from typing import Any
from uuid import uuid4
from zipfile import BadZipFile
import tempfile

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils.exceptions import InvalidFileException
from tqdm import tqdm

from ..database.logger import SQLiteLogger
from ..qr.generator import create_decorated_qr_image
from ..utils.paths import output_excel_path, checkpoint_key
from ..utils.upi import build_upi_deep_link
from ..excel.parsers import find_header_index, parse_amount
from ..excel.operations import setup_qr_column, embed_qr_image, add_qr_hyperlink

from .exceptions import ExcelProcessingError, ConfigurationError, InvalidAmountError
from .models import ProcessConfig, ProcessSummary, QRMode, BillingMode


def _cleanup_temp_files(temp_files: list[Path]) -> None:
    for file in temp_files:
        try:
            file.unlink(missing_ok=True)
        except OSError:
            continue


def _prepare_workbook(input_path: Path, output_path: Path) -> tuple[Workbook, Worksheet, bool]:
    if not input_path.exists():
        raise FileNotFoundError(f"Input file not found: {input_path}")
    if input_path.suffix.lower() != ".xlsx":
        raise ConfigurationError("Input must be an .xlsx file.")

    copied_fresh = False
    if not output_path.exists():
        shutil.copy2(input_path, output_path)
        copied_fresh = True

    try:
        wb = load_workbook(output_path)
    except (InvalidFileException, BadZipFile) as exc:
        raise ExcelProcessingError(f"Cannot open workbook — file may be corrupted: {exc}") from exc

    if not wb.worksheets:
        raise ExcelProcessingError("Workbook has no worksheets.")
    
    return wb, wb.worksheets[0], copied_fresh

def _resolve_headers(ws: Worksheet, config: ProcessConfig) -> tuple[int, int, int | None, int, str]:
    # Scan first 10 rows for the amount header provided by user
    amount_col = None
    header_row = 1
    for r in range(1, 11):
        found_col = find_header_index(ws, config.amount_col_name, row=r)
        if found_col is not None:
            amount_col = found_col
            header_row = r
            break

    if amount_col is None:
        raise ExcelProcessingError(f'Required amount column "{config.amount_col_name}" not found in the first 10 rows.')

    vpa_middle_col: int | None = None
    if config.billing_mode == BillingMode.CUSTOM:
        vpa_middle_col = find_header_index(ws, config.vpa_middle_col_name, row=header_row)
        if vpa_middle_col is None:
            raise ExcelProcessingError(f'VPA middle column "{config.vpa_middle_col_name}" not found.')

    qr_col = find_header_index(ws, "payment_qr", row=header_row)
    if qr_col is None:
        # Find the last non-empty column in the header row to determine where to add QR
        last_data_col = 0
        for col in range(1, ws.max_column + 1):
            if ws.cell(row=header_row, column=col).value is not None:
                last_data_col = col
        qr_col = last_data_col + 1
    
    qr_col_letter = setup_qr_column(ws, qr_col, header_row=header_row)

    return header_row, amount_col, vpa_middle_col, qr_col, qr_col_letter


def _process_single_row(
    row_idx: int,
    ws: Worksheet,
    config: ProcessConfig,
    amount_col: int,
    vpa_middle_col: int | None,
    qr_col: int,
    qr_col_letter: str,
    logger: SQLiteLogger,
    temp_dir: Path,
    hyper_dir: Path,
    temp_files: list[Path]
) -> None:
    raw_amount = ws.cell(row=row_idx, column=amount_col).value
    parsed_amount = parse_amount(raw_amount)
    if parsed_amount is None or parsed_amount <= 0:
        logger.log_step(
            row_index=row_idx, amount=None, txn_id=None, step="validate_amount", status="failed",
            error_type="InvalidAmountError", error_message="Amount is missing, non-numeric, zero, or negative."
        )
        raise InvalidAmountError("Invalid amount.")

    if config.billing_mode == BillingMode.CUSTOM and vpa_middle_col:
        raw_middle = ws.cell(row=row_idx, column=vpa_middle_col).value
        middle = str(raw_middle).strip() if raw_middle is not None else ""
        row_vpa = f"{config.vpa_prefix}{middle}{config.vpa_suffix}"
    else:
        row_vpa = config.vpa

    txn_id = str(uuid4())
    current_step = "generate_upi_url"
    
    upi_url = build_upi_deep_link(
        vpa=row_vpa,
        payee_name=config.payee_name,
        amount=parsed_amount,
        txn_id=txn_id,
        note=config.note,
    )
    logger.log_step(row_index=row_idx, amount=parsed_amount, txn_id=txn_id, step=current_step, status="success")

    current_step = "create_qr_with_logo"
    qr_image = create_decorated_qr_image(upi_url, logo_path=config.logo_path)
    logger.log_step(row_index=row_idx, amount=parsed_amount, txn_id=txn_id, step=current_step, status="success")

    if config.mode == QRMode.EMBED:
        current_step = "embed_image"
        temp_path = temp_dir / f"row_{row_idx}_{txn_id}.png"
        qr_image.save(temp_path, format="PNG")
        embed_qr_image(ws, temp_path, row_idx, qr_col_letter)
        temp_files.append(temp_path)
    else:
        current_step = "save_image"
        file_name = f"qr_row_{row_idx}.png"
        image_path = hyper_dir / file_name
        qr_image.save(image_path, format="PNG")
        rel_path = f"{hyper_dir.name}/{file_name}"
        add_qr_hyperlink(ws, row_idx, qr_col, rel_path)
        current_step = "add_hyperlink"

    logger.log_step(row_index=row_idx, amount=parsed_amount, txn_id=txn_id, step=current_step, status="success")


def process_workbook(
    input_file: str | Path,
    config: ProcessConfig,
    *,
    stop_after_rows: int | None = None,
) -> ProcessSummary:
    input_path = Path(input_file).expanduser().resolve()
    db_path = Path(config.db_path).expanduser().resolve()
    output_path = output_excel_path(input_path)

    successful = failed = skipped = total_rows = 0
    interrupted = False
    resumed_from = None
    wb = ws = None
    temp_files: list[Path] = []
    temp_dir = output_path.parent / f".{output_path.stem}_tmp_qr"
    hyper_dir = output_path.parent / f"{output_path.stem}_qr_images"

    with SQLiteLogger(db_path) as logger:
        logger.set_session_state(status="running", input_file=input_path, output_file=output_path)
        logger.log_session_event(step="run_started", status="success")

        try:
            wb, ws, copied_fresh = _prepare_workbook(input_path, output_path)
            header_row, amount_col, vpa_middle_col, qr_col, qr_col_letter = _resolve_headers(ws, config)

            run_key = checkpoint_key(input_path, output_path, ws.title)
            last_successful_row = logger.get_checkpoint(run_key)
            if copied_fresh and last_successful_row > 1:
                logger.reset_checkpoint(run_key)
                last_successful_row = 1

            resumed_from = last_successful_row if last_successful_row > 1 else None
            start_row = max(header_row + 1, last_successful_row + 1)
            total_rows = max(ws.max_row - header_row, 0)

            if config.mode == QRMode.EMBED:
                with tempfile.TemporaryDirectory(dir=output_path.parent, prefix=f".{output_path.stem}_tmp_qr_") as td:
                    temp_dir = Path(td)
                    
                    rows_to_process = list(range(start_row, ws.max_row + 1))
                    progress = tqdm(rows_to_process, total=len(rows_to_process), desc="Processing rows", unit="row")

                    try:
                        for row_idx in progress:
                            try:
                                _process_single_row(
                                    row_idx, ws, config, amount_col, vpa_middle_col, qr_col, qr_col_letter,
                                    logger, temp_dir, hyper_dir, temp_files
                                )
                                logger.update_checkpoint(run_key, row_idx)
                                successful += 1

                                if stop_after_rows is not None and successful >= stop_after_rows:
                                    raise KeyboardInterrupt("Simulated interruption for testing.")
                            except InvalidAmountError:
                                skipped += 1
                            except KeyboardInterrupt:
                                raise
                            except Exception as exc:
                                failed += 1
                                logger.log_step(
                                    row_index=row_idx, amount=None, txn_id=None, step="process_row", status="failed",
                                    error_type=type(exc).__name__, error_message=str(exc)
                                )
                    except KeyboardInterrupt:
                        interrupted = True
                        logger.log_session_event(step="processing_interrupted", status="failed", error_message="Execution interrupted.")
                    finally:
                        progress.close()

                    if wb is not None:
                        wb.save(output_path)
                        wb.close()
                        wb = None
            else:
                hyper_dir.mkdir(parents=True, exist_ok=True)
                rows_to_process = list(range(start_row, ws.max_row + 1))
                progress = tqdm(rows_to_process, total=len(rows_to_process), desc="Processing rows", unit="row")

                try:
                    for row_idx in progress:
                        try:
                            _process_single_row(
                                row_idx, ws, config, amount_col, vpa_middle_col, qr_col, qr_col_letter,
                                logger, temp_dir, hyper_dir, temp_files
                            )
                            logger.update_checkpoint(run_key, row_idx)
                            successful += 1

                            if stop_after_rows is not None and successful >= stop_after_rows:
                                raise KeyboardInterrupt("Simulated interruption for testing.")
                        except InvalidAmountError:
                            skipped += 1
                        except KeyboardInterrupt:
                            raise
                        except Exception as exc:
                            failed += 1
                            logger.log_step(
                                row_index=row_idx, amount=None, txn_id=None, step="process_row", status="failed",
                                error_type=type(exc).__name__, error_message=str(exc)
                            )
                except KeyboardInterrupt:
                    interrupted = True
                    logger.log_session_event(step="processing_interrupted", status="failed", error_message="Execution interrupted.")
                finally:
                    progress.close()

                if wb is not None:
                    wb.save(output_path)
                    wb.close()
                    wb = None

            status = "interrupted" if interrupted else ("completed_with_errors" if failed + skipped > 0 else "completed")
            logger.set_session_state(status=status, input_file=input_path, output_file=output_path)
            logger.log_summary(total_rows=total_rows, successful_rows=successful, failed_rows=failed, skipped_rows=skipped, output_file=output_path)
            logger.log_session_event(step="run_finished", status="success")

            return ProcessSummary(
                total_rows=total_rows, successful=successful, failed=failed, skipped=skipped,
                output_file=output_path, resumed_from=resumed_from, interrupted=interrupted,
                status=status, session_id=logger.session_id
            )

        except sqlite3.Error as db_err:
            if wb is not None:
                try:
                    wb.save(output_path)
                    wb.close()
                except Exception:
                    pass
            raise ExcelProcessingError("Database failure during processing.") from db_err
        except Exception as exc:
            if wb is not None:
                try:
                    wb.close()
                except Exception:
                    pass
            logger.log_session_event(step="setup_failed", status="failed", error_type=type(exc).__name__, error_message=str(exc))
            logger.set_session_state(status="setup_failed", input_file=input_path, output_file=output_path)
            return ProcessSummary(
                total_rows=0, successful=0, failed=1, skipped=0, output_file=output_path,
                resumed_from=None, interrupted=False, status="setup_failed", session_id=logger.session_id, error_message=str(exc)
            )
