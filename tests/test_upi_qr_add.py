from __future__ import annotations

import sqlite3
from pathlib import Path

import pytest
from rich.console import Console
from typer.testing import CliRunner
from openpyxl import Workbook, load_workbook
from PIL import Image

import upi_qr_add.cli.app as app_module
from upi_qr_add.core.models import BillingMode, ProcessConfig, ProcessSummary, QRMode
from upi_qr_add.core.processor import process_workbook
from upi_qr_add.database.logger import (
    SQLiteLogger,
    archive_or_delete_completed_sessions,
    init_session_db_from_template,
)
from upi_qr_add.qr.generator import create_decorated_qr_image

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def create_excel(path: Path, headers: list[str], rows: list[list[object]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.append(headers)
    for row in rows:
        ws.append(row)
    wb.save(path)


def make_static_config(tmp_path: Path, **kwargs) -> ProcessConfig:
    defaults = dict(
        billing_mode=BillingMode.STATIC,
        vpa="merchant@okaxis",
        payee_name="Demo Store",
        note="Payment for order",
        mode=QRMode.EMBED,
        db_path=tmp_path / "upi_qr_log.db",
    )
    defaults.update(kwargs)
    return ProcessConfig(**defaults)


def make_custom_config(tmp_path: Path, *, vpa_middle_col: str, **kwargs) -> ProcessConfig:
    defaults = dict(
        billing_mode=BillingMode.CUSTOM,
        vpa_prefix="inv.",
        vpa_suffix="@okaxis",
        vpa_middle_col_name=vpa_middle_col,
        payee_name="Custom Merchant",
        note="Payment for order",
        mode=QRMode.EMBED,
        db_path=tmp_path / "upi_qr_custom_log.db",
    )
    defaults.update(kwargs)
    return ProcessConfig(**defaults)


# ---------------------------------------------------------------------------
# QR generation
# ---------------------------------------------------------------------------

def test_decorated_qr_generation_with_and_without_logo(tmp_path: Path) -> None:
    logo_path = tmp_path / "logo.png"
    Image.new("RGBA", (80, 80), (0, 102, 204, 255)).save(logo_path)

    with_logo = create_decorated_qr_image(
        "upi://pay?pa=a@b&pn=Demo&am=1.00&tr=123&tn=Test&cu=INR",
        logo_path=logo_path,
    )
    assert with_logo.size == (420, 420)
    assert with_logo.mode == "RGB"

    without_logo = create_decorated_qr_image(
        "upi://pay?pa=a@b&pn=Demo&am=2.00&tr=124&tn=Test&cu=INR",
        logo_path=tmp_path / "missing_logo.png",
    )
    assert without_logo.size == (420, 420)


# ---------------------------------------------------------------------------
# Static billing mode
# ---------------------------------------------------------------------------

def test_embed_mode_success_and_invalid_rows(tmp_path: Path) -> None:
    input_file = tmp_path / "orders.xlsx"
    create_excel(
        input_file,
        headers=["OrderID", "Amount"],
        rows=[
            [1, 100],
            [2, None],
            [3, "abc"],
            [4, -1],
            [5, 0],
            [6, "49.50"],
        ],
    )
    config = make_static_config(tmp_path)

    summary = process_workbook(input_file, config)

    assert summary.total_rows == 6
    assert summary.successful == 2
    assert summary.skipped == 4
    assert summary.failed == 0
    assert summary.status == "completed_with_errors"
    assert summary.output_file.exists()

    wb = load_workbook(summary.output_file)
    ws = wb.active
    assert ws.cell(row=1, column=3).value == "payment_qr"
    assert len(ws._images) == 2

    conn = sqlite3.connect(config.db_path)
    steps = conn.execute("SELECT step, status FROM row_logs").fetchall()
    conn.close()
    assert ("create_qr_with_logo", "success") in steps
    assert ("validate_amount", "failed") in steps


def test_hyperlink_mode_creates_qr_files_and_links(tmp_path: Path) -> None:
    input_file = tmp_path / "sales.xlsx"
    create_excel(
        input_file,
        headers=["Amount"],
        rows=[
            [10],
            [20.5],
        ],
    )
    config = make_static_config(
        tmp_path,
        vpa="merchant@okaxis",
        payee_name="Shop Name",
        mode=QRMode.HYPERLINK,
        db_path=tmp_path / "hyper_log.db",
    )

    summary = process_workbook(input_file, config)
    assert summary.successful == 2
    assert summary.failed_total == 0
    assert summary.status == "completed"

    wb = load_workbook(summary.output_file)
    ws = wb.active
    link_cell_1 = ws.cell(row=2, column=2)
    link_cell_2 = ws.cell(row=3, column=2)
    assert link_cell_1.value == "Open QR"
    assert link_cell_1.hyperlink is not None
    assert link_cell_2.hyperlink is not None

    image_dir = tmp_path / "sales_with_qr_qr_images"
    assert image_dir.exists()
    png_files = list(image_dir.glob("*.png"))
    assert len(png_files) == 2


def test_missing_amount_column_returns_setup_failed_and_logs(tmp_path: Path) -> None:
    input_file = tmp_path / "missing_amount.xlsx"
    create_excel(
        input_file,
        headers=["Total", "OrderID"],
        rows=[[100, 1]],
    )
    config = make_static_config(tmp_path, db_path=tmp_path / "missing_col.db")

    summary = process_workbook(input_file, config)
    assert summary.status == "setup_failed"
    assert summary.error_message is not None
    assert "Missing required amount column" in summary.error_message

    with SQLiteLogger(config.db_path) as logger:
        events = logger.get_session_events()
        assert any(event["step"] == "setup_failed" for event in events)


def test_resume_after_interruption_uses_checkpoint(tmp_path: Path) -> None:
    input_file = tmp_path / "resume.xlsx"
    create_excel(
        input_file,
        headers=["Amount"],
        rows=[
            [100],
            [200],
            [300],
        ],
    )
    config = make_static_config(
        tmp_path,
        vpa="merchant@okaxis",
        payee_name="Resume Merchant",
        db_path=tmp_path / "resume_log.db",
    )

    first_run = process_workbook(input_file, config, stop_after_rows=1)
    assert first_run.interrupted is True
    assert first_run.successful == 1
    assert first_run.status == "interrupted"

    second_run = process_workbook(input_file, config)
    assert second_run.interrupted is False
    assert second_run.resumed_from == 2
    assert second_run.successful == 2
    assert second_run.status == "completed"

    wb = load_workbook(second_run.output_file)
    ws = wb.active
    assert len(ws._images) == 3


def test_second_row_header_and_first_sheet_only(tmp_path: Path) -> None:
    input_file = tmp_path / "multi_sheet.xlsx"

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "FirstSheet"
    ws1.append(["Invoice export summary title"])
    ws1.append(["Customer", "Balance_Amount(Rs)"])
    ws1.append(["A", 111.25])
    ws1.append(["B", 222.75])

    ws2 = wb.create_sheet("SecondSheet")
    ws2.append(["Amount"])
    ws2.append([9999])
    wb.save(input_file)

    config = make_static_config(tmp_path, db_path=tmp_path / "first_sheet.db")

    summary = process_workbook(input_file, config)
    assert summary.total_rows == 2
    assert summary.successful == 2
    assert summary.failed_total == 0

    output_wb = load_workbook(summary.output_file)
    out_ws1 = output_wb["FirstSheet"]
    out_ws2 = output_wb["SecondSheet"]

    assert out_ws1.cell(row=2, column=3).value == "payment_qr"
    assert len(out_ws1._images) == 2
    assert out_ws2.max_column == 1


# ---------------------------------------------------------------------------
# Custom billing mode
# ---------------------------------------------------------------------------

def test_custom_billing_mode_generation(tmp_path: Path) -> None:
    input_file = tmp_path / "custom_billing.xlsx"
    create_excel(
        input_file,
        headers=["Amount", "UpiMiddle", "CustomerName"],
        rows=[
            [150, "abc123", "Alice"],
            [250, "xyz789", "Bob"],
        ],
    )
    config = make_custom_config(
        tmp_path,
        vpa_middle_col="UpiMiddle",
        payee_name="Shared Payee",
    )

    summary = process_workbook(input_file, config)

    assert summary.total_rows == 2
    assert summary.successful == 2
    assert summary.failed_total == 0
    assert summary.status == "completed"
    assert summary.output_file.exists()


def test_custom_billing_invalid_amount_rows_are_skipped(tmp_path: Path) -> None:
    input_file = tmp_path / "custom_skip.xlsx"
    create_excel(
        input_file,
        headers=["Amount", "UpiMiddle", "CustomerName"],
        rows=[
            [300, "mid1", "Carol"],
            [None, "mid2", "Dave"],
            [-50, "mid3", "Eve"],
        ],
    )
    config = make_custom_config(
        tmp_path,
        vpa_middle_col="UpiMiddle",
        db_path=tmp_path / "custom_skip_log.db",
    )

    summary = process_workbook(input_file, config)
    assert summary.successful == 1
    assert summary.skipped == 2
    assert summary.status == "completed_with_errors"


def test_custom_billing_missing_vpa_column_raises_setup_failed(tmp_path: Path) -> None:
    input_file = tmp_path / "bad_col.xlsx"
    create_excel(
        input_file,
        headers=["Amount", "CustomerName"],
        rows=[[100, "Frank"]],
    )
    config = make_custom_config(
        tmp_path,
        vpa_middle_col="NonExistentCol",
        db_path=tmp_path / "bad_col_log.db",
    )

    summary = process_workbook(input_file, config)
    assert summary.status == "setup_failed"
    assert summary.error_message is not None
    assert "NonExistentCol" in summary.error_message


# ---------------------------------------------------------------------------
# Logger lifecycle
# ---------------------------------------------------------------------------

def test_session_db_lifecycle_helpers(tmp_path: Path) -> None:
    template_db = tmp_path / "template.db"
    sqlite3.connect(template_db).close()

    completed_db = init_session_db_from_template(
        "completed_1",
        sessions_dir=tmp_path,
        template_db=template_db,
    )
    interrupted_db = init_session_db_from_template(
        "interrupted_1",
        sessions_dir=tmp_path,
        template_db=template_db,
    )

    with SQLiteLogger(completed_db, session_id="completed_1") as logger:
        logger.set_session_state(status="completed")
    with SQLiteLogger(interrupted_db, session_id="interrupted_1") as logger:
        logger.set_session_state(status="interrupted")

    deleted = archive_or_delete_completed_sessions(sessions_dir=tmp_path)
    assert deleted == 1
    assert not completed_db.exists()
    assert interrupted_db.exists()


# ---------------------------------------------------------------------------
# CLI / UI
# ---------------------------------------------------------------------------

def test_error_viewer_prints_failed_rows(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    import upi_qr_add.cli.display as display_repo
    db_path = tmp_path / "session_errors.db"
    with SQLiteLogger(db_path, session_id="s1") as logger:
        logger.log_step(
            row_index=4,
            amount=10.0,
            txn_id="txn-1",
            step="embed_image",
            status="failed",
            error_type="IOError",
            error_message="disk full",
        )
        logger.set_session_state(status="completed_with_errors")

    fake_console = Console(record=True, width=180)
    monkeypatch.setattr(display_repo, "console", fake_console)
    display_repo.show_last_run_errors(db_path)
    output = fake_console.export_text()
    assert "embed_image" in output
    assert "disk full" in output


def test_main_menu_loop_runs_and_quits(monkeypatch: pytest.MonkeyPatch, tmp_path: Path) -> None:
    choices = iter(["start", "quit"])
    fake_summary_db = tmp_path / "session_test.db"
    sqlite3.connect(fake_summary_db).close()
    fake_summary = ProcessSummary(
        total_rows=1, successful=1, failed=0, skipped=0,
        output_file=tmp_path / "out.xlsx", resumed_from=None, interrupted=False,
        status="completed", session_id="session_test", error_message=None,
    )

    monkeypatch.setattr(app_module, "choose_main_menu", lambda: next(choices))
    monkeypatch.setattr(app_module, "_run_single_session", lambda: (fake_summary_db, fake_summary))
    monkeypatch.setattr(app_module, "render_title", lambda x: None)
    monkeypatch.setattr(app_module, "archive_or_delete_completed_sessions", lambda: 0)
    monkeypatch.setattr(app_module, "find_latest_session_db", lambda: None)
    
    # disable pure print
    monkeypatch.setattr(app_module, "print_raw", lambda x: None)
    monkeypatch.setattr(app_module, "print_summary", lambda x, y: None)

    runner = CliRunner()
    result = runner.invoke(app_module.app, [])
    assert result.exit_code == 0
